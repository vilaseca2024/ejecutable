import re
import os
import datetime
from pathlib import Path
import fitz 
import openpyxl
from PyPDF2 import PdfReader
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, filedialog, END, Variable, messagebox
from tkinter import Toplevel, Label
import time
# ────────────── CONFIGURACIÓN DE ASSETS Y RUTAS ──────────────

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path("assets/")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()

window.title("DIMEx")
window.geometry("450x370")
window.configure(bg="#202020")
window.iconbitmap(relative_to_assets("logo.ico"))

# ────────────── FUNCIONES DE CARGA DE ARCHIVOS ──────────────

def select_output_path():
    selected_files = filedialog.askopenfilenames(
        title="Seleccionar archivos PDF",
        filetypes=[("Archivos PDF", "*.pdf")],
        multiple=True
    )
    output_entry.delete(0, END)
    if selected_files:
        output_entry.insert(0, '; '.join(selected_files))

def select_output_path_7():
    selected_files = filedialog.askopenfilenames(
        title="Seleccionar archivo Excel (mosol)",
        filetypes=[("Archivos Excel", "*.xlsx")],
        multiple=False
    )
    output_entry7.delete(0, END)
    if selected_files:
        output_entry7.insert(0, selected_files[0])

# ────────────── FUNCIONES PARA TOGGLE DEL MOSOL ──────────────

toggle_img = PhotoImage(file=relative_to_assets("togglebtn_off.png"))
onoff = Variable(value="False")

def toggle_on():
    toggle_img.configure(file=relative_to_assets("togglebtn_on.png"))
    toggle_btn.configure(command=toggle_off)
    onoff.set("True")
    output_entry7.place(x=37.0, y=223.0, width=346.0, height=30.0)
    button_7.place(x=402.0, y=230.4, width=16.0, height=15.2)
    canvas.itemconfigure(image_7, state="normal")
    canvas.itemconfigure(entry_bg_7, state="normal")

def toggle_off():
    toggle_img.configure(file=relative_to_assets("togglebtn_off.png"))
    toggle_btn.configure(command=toggle_on)
    onoff.set("False")
    output_entry7.delete(0, END)
    output_entry7.place_forget()
    button_7.place_forget()
    canvas.itemconfigure(image_7, state="hidden")
    canvas.itemconfigure(entry_bg_7, state="hidden")

# ────────────── FUNCIÓN DE PROCESAMIENTO (INTEGRANDO LA LÓGICA) ──────────────

def eval_expr(expr, row_dict):
    toks = re.findall(r'\d+\.\d+|\d+|[A-Za-z0-9\.]+|[+\-*/]', expr)
    vals = []
    for tk in toks:
        if tk in '+-*/':
            vals.append(tk)
        else:
            try:
                vals.append(float(tk))
            except ValueError:
                raw = row_dict.get(tk, 0) or 0
                vals.append(float(raw))
    i = 0
    while i < len(vals):
        if vals[i] in ('*', '/'):
            op = vals[i]
            a, b = vals[i-1], vals[i+1]
            res = a * b if op == '*' else (a / b if b != 0 else 0)
            vals[i-1:i+2] = [res]
            i -= 1
        else:
            i += 1
    result = vals[0]
    i = 1
    while i < len(vals):
        op, b = vals[i], vals[i+1]
        if op == '+':
            result += b
        elif op == '-':
            result -= b
        i += 2
    return result



def extraer_campos_pdf(ruta_pdf):
    doc = fitz.open(ruta_pdf)
    zona_encabezado = fitz.Rect(0, 0, doc[0].rect.width, 60)
    for pagina in doc:
        pagina.add_redact_annot(zona_encabezado, fill=(1, 1, 1))
        pagina.apply_redactions()
        pagina.insert_text(
            point=(zona_encabezado.x0 + 30, zona_encabezado.y0 + 35),
            text="\n Nueva pagina",
            fontsize=10,
            fontname="helv",
            fill=(0, 0, 0),
        )
    ruta_redacted = ruta_pdf.replace('.pdf', '_redacted.pdf')
    doc.save(ruta_redacted)
    doc.close()
    full_text = ""
    with open(ruta_redacted, 'rb') as f:
        reader = PdfReader(f)
        for page in reader.pages:
            txt = page.extract_text()
            txt = re.sub(r'[^\x20-\x7EñÑáéíóúÁÉÍÓÚüÜ\s]', '', txt)

            if txt:
                full_text += txt + "\n"
    pattern = r'([A-Z]\d*(?:\.\d+)?\.)\s*([\s\S]*?)(?=(?:[A-Z]\d*(?:\.\d+)?\.)|$)'
    matches = re.findall(pattern, full_text, re.DOTALL)

    campos = []
    sublevel_ids = {'H8.', 'E1.', 'I2.'}
    for id_actual, bloque in matches:
        lines = [l.strip() for l in bloque.splitlines() if l.strip()]
        if not lines:
            continue
        
        if id_actual in {'J.', 'G.'}:
            campos.append({
                'id': id_actual,
                'titulo': lines[0],
                'valor': lines[1] if len(lines) > 1 else ''
            })
            for line in lines[2:]:
                line_stripped = line.strip()
                if re.match(r'^(Liquidación|Tipo|Sub totales)', line_stripped):
                    continue

                if line_stripped.startswith("GA"):
                    valor = line_stripped.split()[-1]
                    campos.append({
                        'id': f"{id_actual.rstrip('.')}.GA",
                        'titulo': 'GA',
                        'valor': valor
                    })
                    continue
                elif line_stripped.startswith("IVA"):
                    valor = line_stripped.split()[-1]
                    campos.append({
                        'id': f"{id_actual.rstrip('.')}.IVA",
                        'titulo': 'IVA',
                        'valor': valor
                    })
                    continue
                elif line_stripped.startswith("IDHE"):
                    valor = line_stripped.split()[-1]
                    campos.append({
                        'id': f"{id_actual.rstrip('.')}.IDHE",
                        'titulo': 'IDHE',
                        'valor': valor
                    })
                    continue

                if id_actual == 'G.' and line_stripped.startswith('Total tributos a pagar'):
                    parts = line_stripped.split()
                    campos.append({
                        'id': f"{id_actual.rstrip('.')}.TOTAL",
                        'titulo': ' '.join(parts[:-1]),
                        'valor': parts[-1]
                    })
            continue

        if id_actual in sublevel_ids:
            base = id_actual.rstrip('.')
            base_lines, rest = [], []
            found_num = False
            for line in lines:
                if re.match(r'^\d+\s+', line):
                    found_num = True
                (base_lines if not found_num else rest).append(line)
            if base_lines:
                campos.append({
                    'id': id_actual,
                    'titulo': base_lines[0],
                    'valor': ' '.join(base_lines[1:]).strip()
                })

            cur_id = cur_title = cur_val = ''
            last_valid_num = 0
            for line in rest:
                m = re.match(r'^(\d+)\s+(.*)', line)
                if m:
                    num = int(m.group(1))
                    txt = m.group(2).strip()
                    
                    if base == "H8":
                        is_h8_11_1 = (last_valid_num == 11 and num == 1)
                        is_valid_h8 = is_h8_11_1 or (1 <= num <= 14 and (last_valid_num == 0 or num == last_valid_num + 1))
                        
                        if is_valid_h8:
                            if cur_id:
                                campos.append({
                                    'id': cur_id,
                                    'titulo': cur_title,
                                    'valor': cur_val.strip()
                                })
                            if is_h8_11_1:
                                cur_id = f"{base}.11.1"
                            else:
                                cur_id = f"{base}.{num}"
                                last_valid_num = num
                            cur_title = txt
                            cur_val = ''
                        else:
                            cur_val += ' ' + m.group(1) + ' ' + txt
                    else:
                        cur_val += ' ' + line.strip()
                else:
                    line_strip = line.strip()
                    m_invalid = re.match(r'^([A-Z]+\.)\s*=>\s*(.*)', line_strip)
                    if m_invalid:
                        invalid_prefix = m_invalid.group(1)
                        invalid_content = m_invalid.group(2).replace("=>", "").strip()
                        if cur_val and re.search(r'[A-Z]+$', cur_val):
                            cur_val = cur_val.rstrip() + invalid_prefix + ' ' + invalid_content
                        else:
                            cur_val += ' ' + invalid_prefix + ' ' + invalid_content
                    else:
                        cur_val += ' ' + line_strip

            if cur_id:
                campos.append({
                    'id': cur_id,
                    'titulo': cur_title,
                    'valor': cur_val.strip()
                })

            continue

        titulo = lines[0]
        valor = ' '.join(lines[1:]).strip() if len(lines) > 1 else ''
        campos.append({'id': id_actual, 'titulo': titulo, 'valor': valor})

    for c in campos:
        if c['valor']:
            m = re.match(r'^([^\d,]*\([\w\s]+\))\s+([\d,\.]+)$', c['valor'])
            if m:
                c['titulo'] += ' ' + m.group(1).strip()
                c['valor'] = m.group(2).strip()
    clean = []
    prev = None
    for c in campos:
        if c['id'] == 'A.' and prev and prev['id'] in {'B1.', 'B2.'}:
            prev['valor'] = (prev['valor'] + ' ' + c['titulo'] + ' ' + c['valor']).strip()
        else:
            clean.append(c)
            prev = c
    campos = clean

    
    for c in campos:
  

        if c['id'] == 'F4.':
            c['valor'] = c['valor'].replace(',', '.')
        if c['id'] == 'J.':    
            if not c['valor'].startswith('OI'):
                c['valor'] = ''
      
    for idx, c in enumerate(campos):
        if 'Nueva pagina' in c['valor']:
            if idx >= 4 and all(campos[idx - k - 1]['valor'].strip() == '' for k in range(3)):
                limpia = c['valor'].replace('Nueva pagina', '').strip()
                raw_tokens = [t.strip() for t in re.split(r'\s+', limpia)]
                tokens_clean = [t.rstrip(',') for t in raw_tokens if t]
                num_pat = r'^\d+(?:\.\d+)?$'
                if tokens_clean and all(re.match(num_pat, t) for t in tokens_clean):
                    tokens = tokens_clean
                else:
                    desc, nums, trail = [], [], []
                    for t in tokens_clean:
                        if re.match(num_pat, t):
                            nums.append(t)
                        else:
                            if not nums:
                                desc.append(t)
                            else:
                                trail.append(t)
                    tokens = []
                    if desc:
                        tokens.append(' '.join(desc))
                    tokens.extend(nums)
                    if trail:
                        tokens.append(' '.join(trail))
                if len(tokens) == 5 and all(re.match(r'^\d+(?:\.\d+)?$', t) for t in tokens):
                    for j in range(5):
                        campos[idx - 4 + j]['valor'] = tokens[j]
                else:
                    id_actual = c['id']
                    molde_encontrado = False
                    for ref_idx in range(idx - 1, -1, -1):
                        if campos[ref_idx]['id'] == id_actual and ref_idx >= 4:
                            molde = campos[ref_idx - 4:ref_idx + 1]
                            molde_encontrado = True
                            break
                    if not molde_encontrado:
                        for ref_idx in range(idx + 1, len(campos) - 4):
                            if campos[ref_idx]['id'] == id_actual:
                                molde = campos[ref_idx - 4:ref_idx + 1]
                                molde_encontrado = True
                                break
                    if molde_encontrado:
                        nuevo_bloque = []
                        token_idx = 0
                        for campo_molde in molde:
                            if campo_molde['valor'].strip() != '':
                                if token_idx < len(tokens):
                                    nuevo_bloque.append(tokens[token_idx])
                                    token_idx += 1
                                else:
                                    nuevo_bloque.append('')
                            else:
                                nuevo_bloque.append('')
                        for j in range(5):
                            campos[idx - 4 + j]['valor'] = nuevo_bloque[j]
    
    aux = []
    for c in campos:
       
        if c['id'] in {'B1.', 'B2.'}:
            nuevo_titulo = re.sub(r'^(Importador:|Declarante:)\s*', '', c['titulo'])
            tokens = nuevo_titulo.split()
            doc_tipo = tokens[0] if len(tokens) > 0 else ""
            doc_numero = tokens[1] if len(tokens) > 1 else ""
            nombre = " ".join(tokens[2:]) if len(tokens) > 2 else ""
            m_val = re.search(r'^(.*?)(OEA)(.*)$', c['valor'])
            if m_val:
                parte1 = m_val.group(1).strip()  # Para B2. se usará en el nombre
                categoria = m_val.group(2).strip()
                domicilio = m_val.group(3).strip()
            else:
                categoria = ""
                domicilio = c['valor']
            if c['id'] == 'B2.' and m_val and parte1:
                nombre = nombre + " S.A. " + parte1
            aux.append({'id': c['id'][:-1] + ".TIPO", 'titulo': "Tipo de Documento", 'valor': doc_tipo})
            aux.append({'id': c['id'][:-1] + ".NUMERO", 'titulo': "Nro. de documento", 'valor': doc_numero})
            aux.append({'id': c['id'][:-1] + ".NOMBRE", 'titulo': "Nombre/Razón social", 'valor': nombre})
            aux.append({'id': c['id'][:-1] + ".CATEGORIA", 'titulo': "Categoría", 'valor': categoria})
            aux.append({'id': c['id'][:-1] + ".DOMICILIO", 'titulo': "Domicilio", 'valor': domicilio})
            continue

  
        if c['id'] == 'E1.':
            if ':' in c['titulo']:
                parts = c['titulo'].split(":", 1)
                c['titulo'] = parts[0].strip()
                c['valor'] = (parts[1].strip() + " " + c['valor']).strip()
            aux.append(c)
            continue
        if c['id'] == 'H5.':
            if c['valor'].strip().startswith("Descripción arancelaria:"):
                extra = c['valor'].replace("Descripción arancelaria:", "", 1).strip()
                c['valor'] = ""
                aux.append(c)
                aux.append({'id': "Arancel", 'titulo': "Descripción arancelaria", 'valor': extra})
            else:
                aux.append(c)
            continue
        if c['id'] == 'E14.':
            c['valor'] = c['valor'].replace("Valores y costos", "").strip()
            aux.append(c)
            continue

        aux.append(c)
    campos = aux
    
    nuevos_campos = []
    for c in campos:
        nuevos_campos.append(c)
        if c['id'] == 'L1.':
            m = re.search(
                r'(?:AEREO|EMBARQUE MARÍTIMO|CARRETERA \(CRT\))\s*([A-Z0-9/-]+)',
                c['valor']
            )
            if m:
                codigo = m.group(1)
                nuevos_campos.insert(0, {
                    'id': 'A0.',
                    'titulo': 'Embarque Codigo',
                    'valor': codigo
                })
                
    campos = nuevos_campos

    campos_corregidos = []
    h8_indices = []

    for i, c in enumerate(campos):
        if re.match(r'^H8\.\d+$', c['id']):
            h8_indices.append(i)

    ultimo_h8_idx_real = h8_indices[-1] if h8_indices else None
    ultimo_h8_idx_corregido = None
    en_h8 = False  

    for i, c in enumerate(campos):
        if re.match(r'^H8\.([0-9]|10)$', c['id']):
            if i == ultimo_h8_idx_real:
                campos_corregidos.append(c)
                ultimo_h8_idx_corregido = None
                en_h8 = False
            else:
                ultimo_h8_idx_corregido = len(campos_corregidos)
                campos_corregidos.append(c)
                en_h8 = True
        elif re.match(r'^[A-Z]\.$', c['id']) and en_h8 and ultimo_h8_idx_corregido is not None:
            campos_corregidos[ultimo_h8_idx_corregido]['valor'] += (
                c['id'] + ' ' + (c['titulo'] + ' ' + c['valor']).strip()
            )
        else:
            campos_corregidos.append(c)
            en_h8 = False

    campos = campos_corregidos

    for i, c in enumerate(campos):
        if c['id'] == 'E1.' and (i + 1) < len(campos):
            next_field = campos[i + 1]
            if next_field['id'] != 'E1.1' and re.match(r'^[A-Z]\.$', next_field['id']):
                c['valor'] = c['valor'].strip()+ next_field['id'].strip()  + next_field['titulo'].strip() + " " + next_field['valor'].strip()

    try:
        os.remove(ruta_redacted)
    except:
        pass

    with open("validar.txt", "w", encoding="utf-8") as out:
        for c in campos:
            c['valor'] = c['valor'].replace('Nueva pagina', '')
           
            out.write(f"{c['id']} => {c['titulo']}  =>  {c['valor']}\n")

    return campos


def process_all():
    generate_btn.configure(state="disabled")
    waiting = Toplevel(window)
    waiting.title("Procesando...")
    waiting.geometry("300x100")
    waiting.configure(bg="#202020")
    waiting.transient(window)
    waiting.grab_set()  
    Label(waiting, text="Extrayendo datos, espere...", bg="#202020", fg="#FFFFFF", font=("Roboto", 12)).pack(expand=True, fill="both", padx=20, pady=20)
    waiting.update()  
    pdfs_str = output_entry.get().strip()
    if not pdfs_str:
        messagebox.showwarning("Advertencia", "Debes seleccionar al menos un archivo PDF.")
        generate_btn.configure(state="normal")
        waiting.destroy()
        return
    rutas_pdfs = [ruta.strip() for ruta in pdfs_str.split(';') if ruta.strip()]
    if onoff.get() == "True":
        mosol_path = output_entry7.get().strip()
        if not mosol_path:
            messagebox.showwarning("Advertencia", "El interruptor está activado, pero no se ha seleccionado ningún archivo Excel.")
            generate_btn.configure(state="normal")
            waiting.destroy()
            return
    else:
        mosol_path = str(OUTPUT_PATH / "mosol"/"mosol.xlsx")
        if not os.path.exists(mosol_path):
            messagebox.showerror("Error", f"No se encontró el archivo mosol.xlsx en: {OUTPUT_PATH}")
            generate_btn.configure(state="normal")
            waiting.destroy()
            return
    try:
        wb = openpyxl.load_workbook(mosol_path)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el Excel: {e}")
        generate_btn.configure(state="normal")
        waiting.destroy()
        return

    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    formula_cols = {
        idx: str(h).strip()
        for idx, h in enumerate(headers, start=1)
        if h and any(op in str(h) for op in '+-*/')
    }
    
    current_row = 3 
    for ruta in rutas_pdfs:
        try:
            campos = extraer_campos_pdf(ruta)
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el PDF {ruta}:\n{e}")
            continue
        
        rows_data = []
        resumen = {}
        item_corr = None

        for c in campos:
            id_, val = c['id'], c['valor']
            if not id_.startswith('H') and not id_.startswith('J') and not id_.startswith('I'):
                resumen[id_] = val
                continue
            if id_ == 'H1.':
                if item_corr is not None:
                    rows_data.append(item_corr)
                item_corr = resumen.copy()
            if item_corr is not None:
                item_corr[id_] = val
        if item_corr is not None:
            rows_data.append(item_corr)

        for row_dict in rows_data:
            for col_idx, h in enumerate(headers, start=1):
                if not h:
                    continue
                key = str(h).strip()
                if col_idx in formula_cols:
                    try:
                        v = eval_expr(formula_cols[col_idx], row_dict)
                    except Exception:
                        v = row_dict.get(key, '')
                else:
                    v = row_dict.get(key, '')
                try:
                    ws.cell(row=current_row, column=col_idx, value=float(v))
                except:
                    ws.cell(row=current_row, column=col_idx, value=v)
            current_row += 1

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if save_path:
        try:
            wb.save(save_path)
            messagebox.showinfo("Éxito", f"Guardado en: {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
    else:
        messagebox.showwarning("Cancelado", "No se guardó el archivo.")
    
    generate_btn.configure(state="normal")
    waiting.destroy()

# ────────────── FUNCIONES PARA BOTONES "ABOUT" Y "SETTINGS" ──────────────

def GENERATE():
    generate_btn.configure(state="disabled")
    generate_btn.configure(state="normal")


def handle_btn_press(option):
    if option == "about":
        instrucciones = (
            "1. Para extraer información de una DIM solo debes cargar un archivo y presionar 'Generar Reporte Excel'.\n\n"
            "2. En caso de cambiar el informe a uno que tenga tu propio formato, presiona el botón de engranaje y podrás editar el archivo mosol predefinido.\n\n"
            "3. Si no cambias el archivo mosol y solo generas de una vez ese tipo de informe, activa la opción de cargar mosol DIM y carga un documento Excel con las cabeceras de los IDs en la primera fila.\n\n"
            "4. Para evitar errores no cargues un pdf abierto con Adobe Acrobat.\n\n"
            "5. Espero que el programa te sea útil; ¡éxito! Y cualquier mejora, anótala y házmela saber."
        )
        messagebox.showinfo("Instrucciones", instrucciones)
        
    elif option == "settings":
     
        folder_path = os.path.join(OUTPUT_PATH, "mosol")
        try:
            os.startfile(folder_path)  
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta: {e}")
            

    # elif option == "settings":
    #     # Construir la ruta del archivo mosol.xlsx en el mismo directorio que el script
    #     mosol_path = os.path.join(str(OUTPUT_PATH), "mosol.xlsx")
    #     if os.path.exists(mosol_path) and mosol_path.lower().endswith(".xlsx"):
    #         try:
    #             # Abre directamente el archivo .xlsx con el programa asociado (por ejemplo Excel)
    #             os.startfile(mosol_path)  # Solo disponible en Windows
    #         except Exception as e:
    #             messagebox.showerror("Error", f"No se pudo abrir el archivo mosol.xlsx: {e}")
    #     else:
    #         messagebox.showwarning("Advertencia", f"No se encontró un archivo .xlsx en: {OUTPUT_PATH}")
           

    elif option == "generate":
        try:
            start_time = time.time()   
            process_all()
            end_time = time.time()    

            elapsed_time = end_time - start_time
            print(f"Tiempo de ejecución : {elapsed_time:.2f} segundos")

            generate_button_image.configure(file=relative_to_assets("button_1.png"))
        except Exception as e:
            messagebox.showerror("Error", str(e))
            generate_button_image.configure(file=relative_to_assets("button_1.png"))
            generate_btn.configure(state="normal")

# ────────────── INTERFAZ GRÁFICA (USANDO Tkinter Y LOS ASSETS) ──────────────

canvas = Canvas(
    window,
    bg="#202020",
    height=600,
    width=450,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)
canvas.place(x=0, y=0)

button_image_3 = PhotoImage(file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    activebackground="#202020",
    cursor="heart",
    command=lambda: handle_btn_press("about"),
    relief="flat"
)
button_3.place(x=20.0, y=21.0, width=30.0, height=30.0)
image_image_6 = PhotoImage(file=relative_to_assets("image_6.png"))
canvas.create_image(225.0, 37.0, image=image_image_6)

button_image_4 = PhotoImage(file=relative_to_assets("button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    activebackground="#202020",
    cursor="pirate",
    command=lambda: handle_btn_press("settings"),
    relief="flat"
)
button_4.place(x=400.0, y=21.0, width=30.0, height=30.0)
image_image_5 = PhotoImage(file=relative_to_assets("image_5.png"))
canvas.create_image(224.5, 137.5, image=image_image_5)
entry_image_1 = PhotoImage(file=relative_to_assets("entry_1.png"))
canvas.create_image(210.0, 138.0, image=entry_image_1)
output_entry = Entry(canvas, bd=0, bg="#2D2D2D", fg="#FFFFFF", highlightthickness=0)
output_entry.place(x=37.0, y=123.0, width=346.0, height=30.0)

button_image_2 = PhotoImage(file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=select_output_path,
    activebackground="#202020",
    cursor="target",
    relief="flat"
)
button_2.place(x=402.0, y=130.4, width=16.0, height=15.2)

canvas.create_text(
    20.0,
    98.0,
    anchor="nw",
    text="Ingresa un documento DIM",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    75.0,
    178.0,
    anchor="nw",
    text="¿Desea cargar un mosol de DIM?",
    fill="#FFFFFF",
    font=("Roboto Regular", 14 * -1)
)

toggle_btn = Button(
    image=toggle_img,
    activebackground="#202020",
    relief="flat",
    borderwidth=0,
    highlightthickness=0,
    command=toggle_on  
)
toggle_btn.place(x=27.0, y=175.0)

image_image_7 = PhotoImage(file=relative_to_assets("image_5.png"))
image_7 = canvas.create_image(224.5, 237.5, image=image_image_7)
entry_image_7 = PhotoImage(file=relative_to_assets("entry_1.png"))
entry_bg_7 = canvas.create_image(210.0, 238.0, image=entry_image_7)
output_entry7 = Entry(canvas, bd=0, bg="#2D2D2D", fg="#FFFFFF", highlightthickness=0)
output_entry7.place_forget()
button_image_7 = PhotoImage(file=relative_to_assets("button_2.png"))
button_7 = Button(
    image=button_image_7,
    borderwidth=0,
    highlightthickness=0,
    command=select_output_path_7,
    activebackground="#202020",
    cursor="target",
    relief="flat"
)
button_7.place_forget()
canvas.itemconfigure(image_7, state="hidden")
canvas.itemconfigure(entry_bg_7, state="hidden")

generate_button_image = PhotoImage(file=relative_to_assets("button_1.png"))
generate_btn = Button(
    image=generate_button_image,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: handle_btn_press("generate"),
    activebackground="#202020",
    relief="flat"
)
generate_btn.place(x=18.0, y=300.0, width=414.0, height=47.0)

window.resizable(False, False)
window.mainloop()
